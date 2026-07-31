"""
================================================================================
Excel API 명세서 → Markdown 변환 스크립트
================================================================================
KB증권 OpenAPI 명세서(xlsx)를 AI가 이해하기 쉬운 Markdown 파일로 변환합니다.

--------------------------------------------------------------------------------
사용법 요약
--------------------------------------------------------------------------------
  uv run python -m manage.generate.convert_xlsx_to_md "파일.xlsx"           # 단일 파일 변환
  uv run python -m manage.generate.convert_xlsx_to_md "폴더/"               # 폴더 일괄 변환
  uv run python -m manage.generate.convert_xlsx_to_md "폴더/" --recursive   # 하위 폴더 포함 변환
  uv run python -m manage.generate.convert_xlsx_to_md "폴더/" "출력폴더/"   # 출력 폴더 지정
  uv run python -m manage.generate.convert_xlsx_to_md --help                # 도움말

--------------------------------------------------------------------------------
사용법 상세
--------------------------------------------------------------------------------

1. 단일 파일 변환
   ---------------
   특정 xlsx 파일 하나를 markdown으로 변환합니다.

   기본 사용법:
     uv run python -m manage.generate.convert_xlsx_to_md <xlsx_파일_경로>

   출력 경로 지정:
     uv run python -m manage.generate.convert_xlsx_to_md <xlsx_파일_경로> <출력_md_경로>

   예시:
     uv run python -m manage.generate.convert_xlsx_to_md "명세/GSA10020-체결-20260709-155752.xlsx"
     → 결과: 명세/GSA10020-체결-20260709-155752.md (같은 폴더에 생성)

     uv run python -m manage.generate.convert_xlsx_to_md "명세/GSA10020-체결-20260709-155752.xlsx" "output/체결API.md"
     → 결과: output/체결API.md (지정한 경로에 생성)

2. 폴더 일괄 변환
   ---------------
   폴더 내 모든 xlsx 파일을 한번에 markdown으로 변환합니다.

   기본 사용법:
     uv run python -m manage.generate.convert_xlsx_to_md <폴더_경로>/

   출력 폴더 지정:
     uv run python -m manage.generate.convert_xlsx_to_md <폴더_경로>/ <출력_폴더_경로>/

   예시:
     uv run python -m manage.generate.convert_xlsx_to_md "명세/"
     → 결과: 명세/ 폴더 내 모든 xlsx가 같은 폴더에 md로 변환

     uv run python -m manage.generate.convert_xlsx_to_md "명세/" "output/"
     → 결과: 명세/ 폴더 내 모든 xlsx가 output/ 폴더에 md로 변환

   주의: 폴더 경로는 반드시 '/'로 끝나야 합니다.

3. 하위 폴더 포함 일괄 변환
   ------------------------
   폴더 및 하위 폴더의 모든 xlsx 파일을 변환합니다.

   사용법:
     uv run python -m manage.generate.convert_xlsx_to_md <폴더_경로>/ --recursive

   예시:
     uv run python -m manage.generate.convert_xlsx_to_md "명세/" --recursive
     → 결과: 명세/ 및 하위 폴더의 모든 xlsx를 각각 같은 위치에 md로 변환

--------------------------------------------------------------------------------
출력 형식
--------------------------------------------------------------------------------
- 파일명: 원본과 동일 (확장자만 .md로 변경)
- 인코딩: UTF-8
- 내용 구조:
  1. API 기본 정보 (코드, 명칭, 환경 URL 등)
  2. INPUT (요청 파라미터) 테이블 + JSON 예시
  3. OUTPUT (응답 데이터) 테이블 + JSON 예시
  4. AI 참고 사항 (필드 타입 설명 등)

--------------------------------------------------------------------------------
요구사항
--------------------------------------------------------------------------------
- Python 3.8 이상
- openpyxl 라이브러리 (pip install openpyxl)

================================================================================
"""

import glob as glob_module
import json
import os
import re
import sys

from openpyxl import load_workbook


def parse_api_spec(sheet_data, sheet_name):
    """KB증권 API 명세서 구조를 파싱"""
    result = {
        "sheetName": sheet_name,
        "apiInfo": {},
        "inputFields": [],
        "inputExample": None,
        "outputFields": [],
        "outputExample": None,
    }

    current_section = None
    headers = None

    for row in sheet_data:
        # B열부터 시작 (인덱스 1)
        values = [str(v).strip() if v is not None else "" for v in row[1:]]

        # 빈 행 건너뛰기
        if all(v == "" for v in values):
            continue

        first_val = values[0] if values else ""
        second_val = values[1] if len(values) > 1 else ""

        # INPUT/OUTPUT 섹션 감지
        if re.match(r"^INPUT$", first_val, re.IGNORECASE):
            current_section = "INPUT"
            headers = None
            continue
        if re.match(r"^OUTPUT$", first_val, re.IGNORECASE):
            current_section = "OUTPUT"
            headers = None
            continue

        # JSON 예시 감지
        if first_val.startswith("{"):
            try:
                parsed = json.loads(first_val)
                if current_section == "INPUT":
                    result["inputExample"] = parsed
                elif current_section == "OUTPUT":
                    result["outputExample"] = parsed
            except json.JSONDecodeError:
                pass
            continue

        # 헤더 행 감지
        if re.search(r"항목영문명|필드명|Field", first_val, re.IGNORECASE):
            headers = values
            continue

        # 메타데이터 영역 (INPUT 전)
        if current_section is None:
            if first_val and second_val:
                result["apiInfo"][first_val] = second_val
            if len(values) > 3 and values[2] and values[3]:
                result["apiInfo"][values[2]] = values[3]
            continue

        # 필드 데이터 행
        if headers and first_val and not first_val.startswith("{"):
            field = {}
            for j, header_name in enumerate(headers):
                if j < len(values) and header_name:
                    field[header_name] = values[j] if values[j] else ""

            # 최소 하나의 유효한 값이 있는 경우만 추가
            if any(v != "" for v in field.values()):
                if current_section == "INPUT":
                    result["inputFields"].append(field)
                elif current_section == "OUTPUT":
                    result["outputFields"].append(field)

    return result


def generate_markdown(parsed_spec, filename):
    """AI 친화적인 Markdown 생성"""
    lines = []
    base_name = os.path.splitext(os.path.basename(filename))[0]
    api_match = re.match(r"^([A-Z0-9]+)-(.+?)-\d{8}", base_name)

    lines.append(f"# API 명세서: {base_name}")
    lines.append("")

    # API 기본 정보
    lines.append("## 기본 정보")
    lines.append("")
    lines.append("| 항목 | 값 |")
    lines.append("| --- | --- |")

    if api_match:
        lines.append(f"| API 코드 | `{api_match.group(1)}` |")
        lines.append(f"| API 명 | {api_match.group(2)} |")

    api_info = parsed_spec["apiInfo"]
    for key, value in api_info.items():
        if key and value and value != key:
            display_value = value[:100] + "..." if len(value) > 100 else value
            lines.append(f"| {key} | {display_value} |")
    lines.append("")

    # INPUT 섹션
    if parsed_spec["inputFields"]:
        lines.append("---")
        lines.append("")
        lines.append("## INPUT (요청 파라미터)")
        lines.append("")

        # 유효한 헤더만 필터링
        all_headers = list(parsed_spec["inputFields"][0].keys())
        headers = [h for h in all_headers if h and h.strip()]

        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

        for field in parsed_spec["inputFields"]:
            row = [field.get(h, "").replace("|", "\\|") for h in headers]
            lines.append("| " + " | ".join(row) + " |")
        lines.append("")

    # INPUT 예시
    if parsed_spec["inputExample"]:
        lines.append("### 요청 예시")
        lines.append("")
        lines.append("```json")
        lines.append(json.dumps(parsed_spec["inputExample"], indent=2, ensure_ascii=False))
        lines.append("```")
        lines.append("")

    # OUTPUT 섹션
    if parsed_spec["outputFields"]:
        lines.append("---")
        lines.append("")
        lines.append("## OUTPUT (응답 데이터)")
        lines.append("")

        # 유효한 헤더만 필터링
        all_headers = list(parsed_spec["outputFields"][0].keys())
        headers = [h for h in all_headers if h and h.strip()]

        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

        for field in parsed_spec["outputFields"]:
            row = [field.get(h, "").replace("|", "\\|") for h in headers]
            lines.append("| " + " | ".join(row) + " |")
        lines.append("")

    # OUTPUT 예시
    if parsed_spec["outputExample"]:
        lines.append("### 응답 예시")
        lines.append("")
        lines.append("```json")
        lines.append(json.dumps(parsed_spec["outputExample"], indent=2, ensure_ascii=False))
        lines.append("```")
        lines.append("")

    # AI 참고 사항
    lines.append("---")
    lines.append("")
    lines.append("## AI 참고 사항")
    lines.append("")
    lines.append("이 문서는 KB증권 OpenAPI 명세서입니다.")
    lines.append("")
    lines.append("**요청 시 주의사항:**")
    lines.append("- `dataHeader`에 `ipAddr`(IP주소)와 `macAddr`(MAC주소)를 포함해야 합니다.")
    lines.append("- `dataBody`에 필수 파라미터를 모두 포함해야 합니다.")
    lines.append("")
    lines.append("**응답 구조:**")
    lines.append("- `dataHeader`: 결과 코드, 메시지, 처리 시간 등 메타 정보")
    lines.append("- `dataBody`: 실제 응답 데이터 (배열인 경우 `out2` 등의 키로 포함)")
    lines.append("")
    lines.append("**필드 타입 설명:**")
    lines.append("- `C숫자`: 문자열 (숫자는 최대 길이)")
    lines.append("- `N숫자`: 숫자")
    lines.append("- `P숫자`: 소수점 포함 숫자 (숫자는 소수점 자릿수)")
    lines.append("")

    return "\n".join(lines)


def convert_single_file(input_path, output_path=None):
    """단일 xlsx 파일을 markdown으로 변환"""
    if output_path is None:
        output_path = re.sub(r"\.xlsx$", ".md", input_path, flags=re.IGNORECASE)

    print(f"Excel 파일 읽는 중: {input_path}")

    wb = load_workbook(input_path, read_only=True, data_only=True)
    sheets = wb.sheetnames

    print(f"  시트 {len(sheets)}개 발견")

    sheet_name = sheets[0]
    ws = wb[sheet_name]
    print(f"  시트 처리 중: {sheet_name}")

    # 시트 데이터를 2차원 리스트로 변환
    sheet_data = []
    for row in ws.iter_rows():
        sheet_data.append([cell.value for cell in row])

    wb.close()

    parsed_spec = parse_api_spec(sheet_data, sheet_name)

    print(f"  - API 정보: {len(parsed_spec['apiInfo'])}개 항목")
    print(f"  - INPUT 필드: {len(parsed_spec['inputFields'])}개")
    print(f"  - OUTPUT 필드: {len(parsed_spec['outputFields'])}개")

    print("  Markdown 변환 중...")
    markdown = generate_markdown(parsed_spec, input_path)

    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(markdown)

    print(f"  변환 완료: {output_path}")
    return True


def convert_folder(input_folder, output_folder=None, recursive=False):
    """폴더 내 모든 xlsx 파일을 markdown으로 변환"""
    if output_folder is None:
        output_folder = input_folder

    # xlsx 파일 목록 가져오기
    if recursive:
        pattern = os.path.join(input_folder, "**", "*.xlsx")
        xlsx_files = glob_module.glob(pattern, recursive=True)
    else:
        pattern = os.path.join(input_folder, "*.xlsx")
        xlsx_files = glob_module.glob(pattern)

    if not xlsx_files:
        print(f"변환할 xlsx 파일이 없습니다: {input_folder}")
        return

    print(f"총 {len(xlsx_files)}개 파일 변환 시작")
    print("=" * 60)

    success_count = 0
    fail_count = 0

    for xlsx_file in xlsx_files:
        try:
            # 출력 경로 계산
            if recursive and output_folder != input_folder:
                # 하위 폴더 구조 유지
                rel_path = os.path.relpath(xlsx_file, input_folder)
                output_path = os.path.join(output_folder, rel_path)
                output_path = re.sub(r"\.xlsx$", ".md", output_path, flags=re.IGNORECASE)
            elif output_folder != input_folder:
                # 단일 폴더 변환, 다른 출력 폴더
                filename = os.path.basename(xlsx_file)
                output_path = os.path.join(output_folder, re.sub(r"\.xlsx$", ".md", filename, flags=re.IGNORECASE))
            else:
                # 같은 폴더에 출력
                output_path = re.sub(r"\.xlsx$", ".md", xlsx_file, flags=re.IGNORECASE)

            convert_single_file(xlsx_file, output_path)
            success_count += 1
        except Exception as e:
            print(f"  오류 발생: {xlsx_file} - {e}")
            fail_count += 1

        print("-" * 60)

    print("=" * 60)
    print(f"변환 완료: 성공 {success_count}개, 실패 {fail_count}개")


def print_usage():
    """사용법 출력"""
    print("""
================================================================================
Excel API 명세서 → Markdown 변환 스크립트
================================================================================

사용법:
  1. 단일 파일 변환
     uv run python -m manage.generate.convert_xlsx_to_md <xlsx_파일_경로> [출력_md_경로]

     예시:
       uv run python -m manage.generate.convert_xlsx_to_md "명세/GSA10020-체결.xlsx"
       uv run python -m manage.generate.convert_xlsx_to_md "명세/GSA10020-체결.xlsx" "output/체결.md"

  2. 폴더 일괄 변환 (폴더 경로는 '/'로 끝나야 함)
     uv run python -m manage.generate.convert_xlsx_to_md <폴더_경로>/ [출력_폴더_경로/]

     예시:
       uv run python -m manage.generate.convert_xlsx_to_md "명세/"
       uv run python -m manage.generate.convert_xlsx_to_md "명세/" "output/"

  3. 하위 폴더 포함 일괄 변환
     uv run python -m manage.generate.convert_xlsx_to_md <폴더_경로>/ --recursive

     예시:
       uv run python -m manage.generate.convert_xlsx_to_md "명세/" --recursive

옵션:
  --recursive, -r    하위 폴더 포함 일괄 변환
  --help, -h         도움말 출력
================================================================================
""")


def main():
    args = sys.argv[1:]

    # 도움말 옵션 확인
    if not args or "--help" in args or "-h" in args:
        print_usage()
        sys.exit(0 if "--help" in args or "-h" in args else 1)

    # recursive 옵션 확인
    recursive = "--recursive" in args or "-r" in args
    args = [a for a in args if a not in ("--recursive", "-r")]

    input_path = args[0]
    output_path = args[1] if len(args) > 1 else None

    if not os.path.exists(input_path.rstrip("/").rstrip("\\")):
        print(f"경로를 찾을 수 없습니다: {input_path}")
        sys.exit(1)

    try:
        # 폴더인지 파일인지 판단 (경로가 /로 끝나면 폴더로 처리)
        if input_path.endswith("/") or input_path.endswith("\\") or os.path.isdir(input_path):
            input_folder = input_path.rstrip("/").rstrip("\\")
            output_folder = output_path.rstrip("/").rstrip("\\") if output_path else None
            convert_folder(input_folder, output_folder, recursive)
        else:
            convert_single_file(input_path, output_path)

    except Exception as error:
        print(f"오류 발생: {error}")
        sys.exit(1)


if __name__ == "__main__":
    main()
