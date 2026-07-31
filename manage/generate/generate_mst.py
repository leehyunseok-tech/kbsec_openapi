#!/usr/bin/env python
"""
종목마스터 파이프라인 — mst/origin/*.mst 원본만 갈아 놓으면 문서와 런타임 데이터를 전부 재생성.

실행:
    uv run python -m manage.generate.generate_mst

입력:
    mst/origin/mtsjname.mst / mtsoutjname.mst / FORENMST_US.MST  (KB 배포 원본, UTF-8)
    docs/mst/xlsx/mst_*.xlsx   (KB 공식 필드 명세 — 순번/한글명/비고 코드표)
    docs/api/api-list.md       (프로젝트가 쓰는 API 목록 — 용도에 적힌 API 코드 검증용)

출력:
    docs/mst/xlsx/openapi_mst_*.xlsx   (공식 명세에서 실제 사용/참조 필드만 선별한 표)
    docs/mst/md/openapi_mst_*.md       (위 선별표의 마크다운 판 + 코드표/라벨 변환 규칙)
    mst/api/openapi_field_kospi-kosdaq.mst  (코스피+코스닥 통합, 런타임 데이터)
    mst/api/openapi_field_foren-us.mst      (미국, 런타임 데이터)

원본 → openapi_field_*.mst는 중간 산출물(과거의 openapi_mtsjname.mst 등) 없이 이
스크립트가 한 번에 생성한다. 필드 선별(사용/참조용)과 용도 설명은 이 파일의
CURATION 표가 단일 소스이고, 필드명·코드표는 공식 명세 xlsx에서 읽어온다 —
공식 명세에 없는 순번을 CURATION이 가리키면 즉시 실패하므로, KB가 마스터
레이아웃을 바꾸면 xlsx 명세를 갈아 넣는 순간 어긋난 부분이 드러난다.

⚠️ 필드 인덱스의 근거는 반드시 공식 명세(docs/mst/xlsx/mst_*.xlsx)다.
과거 파이프라인(mst/create_openapi_mst.py + generate_field_reference_mst.py, 폐지됨)은
다른 증권사 마스터 레이아웃을 가정한 라벨을 쓰는 바람에 실제 데이터와 어긋났었다
(예: 코스피 순번12 '현금증거금율구분'(B030=30%)을 '매매수량단위코드(D040=1주)'로,
순번26 '소수점매매상태'(0:정상/3:매수·매도불가)를 '주문유형'(0:불가/3:가능)으로 —
의미가 정반대. 미국 순번25 'ETF 데이터 기준일자'를 '소수점매매가능'으로 해석해
날짜값이 섞여 보였던 것도 같은 원인). 이 스크립트는 공식 명세 기준으로 교정했고,
실데이터 전수 스캔으로 각 필드의 값 분포가 공식 코드표와 일치함을 확인했다.

코드 → 라벨 변환 원칙: 공식 코드표에 있는 값만 한글 라벨로 바꾸고, 없는 값은
추측하지 않고 "{원본코드}(참조표 미등재)"로 남긴다. 예외 두 가지는 아래 표에
주석으로 근거를 달아뒀다(코스피 증권그룹ID 'EN', 미국 매매구분 '0').

해외 거래소코드(NAS/NYS/AMX)는 번역하지 않는다 — buy/sell/srch 명령이 이 값을
KB API의 krx_cd 파라미터로 그대로 전달하므로, 표시용 한글명은 별도 컬럼(거래소명)
으로만 제공한다.
"""

import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent.parent  # sys.path 부트스트랩 전용 — 경로 상수는 src/paths.py에서
sys.path.insert(0, str(ROOT))  # 파일 직접 실행(-m 없이) 시에도 src.paths import가 풀리도록
from src.paths import (
    API_LIST_MD,
)
from src.paths import (
    DOCS_MST_MD_DIR as MD_DIR,
)
from src.paths import (  # noqa: E402
    DOCS_MST_XLSX_DIR as XLSX_DIR,
)
from src.paths import (
    MST_ORIGIN_DIR as ORIGIN_DIR,
)
from src.paths import (
    MST_RUNTIME_DIR as OUT_DIR,
)

# ── 코드 → 라벨 참조표 (근거: docs/mst/xlsx/mst_*.xlsx 비고 컬럼) ────────────

# 코스피 순번4 증권그룹ID. 'EN'은 공식 코드표에 빠져 있지만, 실데이터에서 EN인
# 384건 중 381건의 ETP상품구분코드(순번8)가 3(ETN)이라 ETN으로 확정했다.
SEC_GROUP = {
    "ST": "주식",
    "MF": "증권투자회사",
    "RT": "부동산투자회사",
    "SC": "선박투자회사",
    "IF": "사회간접자본투융자회사",
    "DR": "주식예탁증서",
    "EW": "ELW",
    "EF": "ETF",
    "EN": "ETN",
    "SW": "신주인수권증권",
    "SR": "신주인수권증서",
    "BC": "수익증권",
    "FE": "해외ETF",
    "FS": "외국주권",
    "KN": "코넥스",
    "PF": "상장형수익증권",
    "BI": "BDC투자회사",
    "BB": "BDC수익증권",
}

# 코스피 순번26 / 코스닥 순번21 소수점매매상태
DECIMAL_STATE = {"0": "정상", "1": "매수불가", "2": "매도불가", "3": "매수/매도불가"}

# 미국 순번1 거래소코드 → 표시용 거래소명 (원본 코드는 krx_cd로 그대로 쓰므로 별도 컬럼)
EXCHANGE_NAME = {
    "AMX": "아멕스",
    "NYS": "뉴욕",
    "NAS": "나스닥",
    "HKS": "홍콩",
    "SHS": "상해",
    "SZS": "심천",
    "TSE": "도쿄",
    "HNX": "하노이",
    "HSX": "호치민",
}

# 미국 순번10 종목타입
US_STOCK_TYPE = {
    "1": "주식",
    "2": "DR",
    "3": "미국워런트",
    "4": "미국우선주",
    "7": "ETF/ETN",
    "22": "워런트",
    "34": "지수",
}

# 미국 순번12 매매구분(SELL ONLY 구분). 공식 코드표는 1/2만 정의하지만 "SELL ONLY
# 구분"이라는 필드 성격상 0(전 종목의 실데이터 값)은 제한 없음으로 확정할 수 있다.
US_TRADE_RESTRICTION = {"0": "제한없음", "1": "매수불가(SELL ONLY)", "2": "매매불가"}

# 미국 순번33 소수점매매대상종목여부 (1:대상 2:비대상)
US_DECIMAL_YN = {"1": "소수점매매가능", "2": "소수점매매불가"}


def _label(value, table):
    if value in table:
        return table[value]
    if not value:
        return ""
    return f"{value}(참조표 미등재)"


def _yn(value, yes_label, no_label):
    if value == "Y":
        return yes_label
    if value == "N":
        return no_label
    if not value:
        return ""
    return f"{value}(참조표 미등재)"


def _unit(value):
    """주문/거래 단위 수치('00001', '1' 등) → '1주'. 숫자가 아니면 미등재 처리."""
    if not value:
        return ""
    if value.isdigit():
        return f"{int(value)}주"
    return f"{value}(참조표 미등재)"


# ── 필드 선별표 (사용/참조용) — 이 파이프라인의 단일 소스 ────────────────────
# seq: 공식 명세(mst_*.xlsx)의 순번. cat '사용'인 필드만 openapi_field_*.mst에 실린다.
# usage의 API 코드는 docs/api/api-list.md에 실존하는지 실행 시 검증한다.

CURATION = {
    "kospi": [
        dict(seq=1, cat="사용", usage="주문(SSAM1801/SSAM1802 등)·시세(IVU10140 등)의 is_cd 파라미터로 직접 사용"),
        dict(seq=2, cat="사용", usage="종목 검색(/stcd)·자연어 종목명 해석 용도"),
        dict(seq=4, cat="사용", usage="종목 유형(주식/ETF/ETN 등) 표시 및 필터링"),
        dict(seq=11, cat="사용", usage="주문 수량 단위 확인 (SSAM1802 매수주문 등)"),
        dict(seq=15, cat="사용", usage="주문 전 확인 필수 (Y=주문 불가)"),
        dict(seq=21, cat="사용", usage="주문 전 확인 필수 (관리종목=매매 유의)"),
        dict(seq=25, cat="사용", usage="소수점주문(SSAM5762/SSAM5763) 가능 여부"),
        dict(seq=26, cat="사용", usage="소수점주문(SSAM5762/SSAM5763) 매수/매도 제한 상태"),
        dict(seq=5, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=6, cat="참조용", usage="정리매매 종목 확인 (표시 제외 — 사용자 요청)"),
        dict(seq=7, cat="참조용", usage="KOSPI200 구성종목 필터링"),
        dict(seq=12, cat="참조용", usage="현금 증거금율 확인 (OpenAPI 주문에는 미사용)"),
        dict(seq=13, cat="참조용", usage="신용거래 시 참조 (OpenAPI 미지원)"),
        dict(seq=14, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=19, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=20, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=22, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=23, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=33, cat="참조용", usage="시가총액 기준 필터링"),
    ],
    "kosdaq": [
        dict(seq=1, cat="사용", usage="주문(SSAM1801/SSAM1802 등)·시세(IVU10140 등)의 is_cd 파라미터로 직접 사용"),
        dict(seq=2, cat="사용", usage="종목 검색(/stcd)·자연어 종목명 해석 용도"),
        dict(seq=14, cat="사용", usage="주문 전 확인 필수 (관리종목=매매 유의)"),
        dict(seq=19, cat="사용", usage="주문 전 확인 필수 (Y=주문 불가)"),
        dict(seq=20, cat="사용", usage="소수점주문(SSAM5762/SSAM5763) 가능 여부"),
        dict(seq=21, cat="사용", usage="소수점주문(SSAM5762/SSAM5763) 매수/매도 제한 상태"),
        dict(seq=3, cat="참조용", usage="현금 증거금율 확인 (OpenAPI 주문에는 미사용)"),
        dict(seq=4, cat="참조용", usage="신용거래 시 참조 (OpenAPI 미지원)"),
        dict(seq=5, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=6, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=7, cat="참조용", usage="정리매매 종목 확인 (표시 제외 — 사용자 요청)"),
        dict(seq=8, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=13, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=15, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=16, cat="참조용", usage="투자 주의 종목 필터링"),
        dict(seq=18, cat="참조용", usage="KOSDAQ150 구성종목 필터링"),
        dict(seq=27, cat="참조용", usage="시가총액 기준 필터링"),
    ],
    "us": [
        dict(
            seq=1, cat="사용", usage="주문(SKAM2101)·시세(GSS10030 등)의 krx_cd 파라미터로 원본 그대로 전달 — 번역 금지"
        ),
        dict(seq=2, cat="사용", usage="주문(SKAM2101)·시세(GSS10030 등)의 is_cd 파라미터로 직접 사용 (티커)"),
        dict(seq=3, cat="사용", usage="종목 검색(/stcd)·자연어 종목명 해석 용도"),
        dict(seq=4, cat="사용", usage="종목 검색(영문명) 용도"),
        dict(seq=6, cat="사용", usage="소수점주문(SKAM2201) crncy_ccd 연계"),
        dict(seq=10, cat="사용", usage="종목 유형(주식/ETF 등) 표시 및 필터링"),
        dict(seq=12, cat="사용", usage="주문 전 확인 (SELL ONLY/매매불가 종목 구분)"),
        dict(seq=14, cat="사용", usage="매수 주문 수량 단위 확인 (SKAM2101)"),
        dict(seq=15, cat="사용", usage="매도 주문 수량 단위 확인 (SKAM2101)"),
        dict(seq=33, cat="사용", usage="소수점주문(SKAM2201/SKAM2202) 가능 여부"),
        dict(seq=5, cat="참조용", usage="국가별 종목 필터링"),
        dict(seq=7, cat="참조용", usage="국제증권식별번호 (종목 식별)"),
        dict(seq=9, cat="참조용", usage="가격 소수점 자리수 (표시 포맷)"),
        dict(seq=11, cat="참조용", usage="업종별 종목 필터링"),
        dict(seq=18, cat="참조용", usage="ETF/ETN 필터링"),
        dict(seq=32, cat="참조용", usage="PTP 종목 필터링 (세금 유의)"),
    ],
}

MARKETS = {
    "kospi": dict(
        title="코스피(KOSPI) 종목 마스터 - 주문 관련 사용 필드",
        source_xlsx="mst_코스피_mtsjname.xlsx",
        origin_mst="mtsjname.mst",
        out_xlsx="openapi_mst_코스피_mtsjname.xlsx",
        out_md="openapi_mst_코스피_mtsjname.md",
    ),
    "kosdaq": dict(
        title="코스닥(KOSDAQ) 종목 마스터 - 주문 관련 사용 필드",
        source_xlsx="mst_코스닥_mtsoutjname.xlsx",
        origin_mst="mtsoutjname.mst",
        out_xlsx="openapi_mst_코스닥_mtsoutjname.xlsx",
        out_md="openapi_mst_코스닥_mtsoutjname.md",
    ),
    "us": dict(
        title="해외주식(미국) 종목 마스터 - 주문 관련 사용 필드",
        source_xlsx="mst_해외주식_FORENMST_US.xlsx",
        origin_mst="FORENMST_US.MST",
        out_xlsx="openapi_mst_해외주식_FORENMST_US.xlsx",
        out_md="openapi_mst_해외주식_FORENMST_US.md",
    ),
}


# ── 입력 읽기 ────────────────────────────────────────────────────────────────


def read_official_fields(xlsx_name):
    """공식 명세 xlsx → {순번: {'name', 'note'}}.

    세 파일의 컬럼 구성이 조금씩 다르므로('순번 '으로 시작하는 헤더 행을 찾은 뒤
    '한글명'/'비고' 제목 위치를 사용) 제목 기준으로 읽는다.
    """
    ws = openpyxl.load_workbook(XLSX_DIR / xlsx_name).active
    rows = [[("" if c is None else str(c).strip()) for c in r] for r in ws.iter_rows(values_only=True)]
    header_idx = next(i for i, r in enumerate(rows) if r and r[0].replace(" ", "") == "순번")
    header = [c.replace(" ", "") for c in rows[header_idx]]
    name_col = header.index("한글명")
    note_col = header.index("비고")
    fields = {}
    for r in rows[header_idx + 1 :]:
        if not r or not r[0]:
            continue
        try:
            seq = int(float(r[0]))
        except ValueError:
            continue
        note = re.sub(r"\s*\n\s*", " / ", r[note_col]) if note_col < len(r) else ""
        fields[seq] = {"name": r[name_col].strip(), "note": note}
    if not fields:
        raise RuntimeError(f"{xlsx_name}: 필드 행을 하나도 읽지 못했습니다")
    return fields


def read_api_codes():
    """api-list.md의 표 첫 컬럼에서 API 코드 집합을 추출 (용도 문구 검증용)."""
    text = API_LIST_MD.read_text(encoding="utf-8")
    return set(re.findall(r"^\|\s*([A-Z]{2,4}\d{4,5})\s*\|", text, flags=re.MULTILINE))


def read_origin_mst(filename, min_cols):
    """원본 .mst를 읽어 파이프 분리 행 리스트로 반환. UTF-8(BOM 허용) 우선, cp949 폴백."""
    path = ORIGIN_DIR / filename
    for enc in ("utf-8-sig", "cp949"):
        try:
            text = path.read_text(encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise RuntimeError(f"{path}: utf-8/cp949 어느 쪽으로도 읽을 수 없습니다")
    rows = [line.split("|") for line in text.splitlines() if line.strip()]
    bad = [i for i, r in enumerate(rows) if len(r) < min_cols]
    if bad:
        raise RuntimeError(
            f"{filename}: {len(bad)}개 행의 컬럼 수가 {min_cols} 미만입니다 (첫 예: {bad[0] + 1}행째). "
            f"KB가 마스터 레이아웃을 바꿨을 수 있으니 docs/mst/xlsx/mst_*.xlsx 공식 명세부터 갱신하세요."
        )
    return rows


def validate_curation(api_codes):
    """CURATION의 순번이 공식 명세에 실존하는지, 용도의 API 코드가 api-list에 있는지 확인."""
    officials = {}
    for key, market in MARKETS.items():
        officials[key] = read_official_fields(market["source_xlsx"])
        for item in CURATION[key]:
            if item["seq"] not in officials[key]:
                raise RuntimeError(f"{key}: CURATION 순번 {item['seq']}이 공식 명세 {market['source_xlsx']}에 없습니다")
            for code in re.findall(r"[A-Z]{2,4}\d{4,5}", item["usage"]):
                if code not in api_codes:
                    raise RuntimeError(
                        f"{key} 순번 {item['seq']}: 용도에 적힌 API '{code}'가 docs/api/api-list.md에 없습니다"
                    )
    return officials


# ── 1단계: openapi_mst_*.xlsx ────────────────────────────────────────────────


def write_openapi_xlsx(key, official):
    market = MARKETS[key]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "필드사용현황"
    ws.append([market["title"]])
    ws.append([])
    ws.append(["순번", "구분", "필드명", "설명", "용도"])
    for item in sorted(CURATION[key], key=lambda x: (x["cat"] != "사용", x["seq"])):
        f = official[item["seq"]]
        ws.append([item["seq"], item["cat"], f["name"], f["note"] or "-", item["usage"]])
    path = XLSX_DIR / market["out_xlsx"]
    wb.save(path)
    print(f"생성됨: {path}")


# ── 2단계: openapi_mst_*.md ──────────────────────────────────────────────────

FIELD_MST_NOTES = {
    "kospi": """\
`mst/api/openapi_field_kospi-kosdaq.mst`(코스피+코스닥 통합, UTF-8, 헤더 행 포함)의
코스피 행은 위 ✅ 사용 필드를 아래 컬럼으로 변환해 담는다:

| 컬럼 | 원본 순번 | 변환 |
|------|-----------|------|
| 시장구분 | - | 고정값 `KOSPI` |
| 종목코드 | 1 | 원본 그대로 |
| 종목명 | 2 | 원본 그대로 |
| 종목구분 | 4 | 증권그룹ID 코드표 적용 (ST→주식, EF→ETF, EN→ETN 등) |
| 관리종목여부 | 21 | Y→관리종목, N→정상 |
| 거래정지여부 | 15 | Y→거래정지, N→정상 |
| 매수주문단위 | 11 | 수치→`n주` (예: 00001→1주) |
| 소수점매매가능여부 | 25 | Y→소수점매매가능, N→소수점매매불가 |
| 소수점매매상태 | 26 | 0→정상, 1→매수불가, 2→매도불가, 3→매수/매도불가 |

코드표에 없는 값은 추측하지 않고 `{원본코드}(참조표 미등재)`로 남긴다.
증권그룹ID `EN`은 공식 코드표에 빠져 있지만 실데이터의 ETP상품구분코드(순번8)가
3(ETN)인 점을 근거로 ETN으로 확정했다.""",
    "kosdaq": """\
`mst/api/openapi_field_kospi-kosdaq.mst`(코스피+코스닥 통합, UTF-8, 헤더 행 포함)의
코스닥 행은 위 ✅ 사용 필드를 아래 컬럼으로 변환해 담는다:

| 컬럼 | 원본 순번 | 변환 |
|------|-----------|------|
| 시장구분 | - | 고정값 `KOSDAQ` |
| 종목코드 | 1 | 원본 그대로 |
| 종목명 | 2 | 원본 그대로 |
| 종목구분 | - | 공란 (코스닥 원본에는 증권그룹ID 필드가 없음) |
| 관리종목여부 | 14 | Y→관리종목, N→정상 |
| 거래정지여부 | 19 | Y→거래정지, N→정상 |
| 매수주문단위 | - | 공란 (코스닥 원본에는 주문단위 필드가 없음) |
| 소수점매매가능여부 | 20 | Y→소수점매매가능, N→소수점매매불가 |
| 소수점매매상태 | 21 | 0→정상, 1→매수불가, 2→매도불가, 3→매수/매도불가 |

코드표에 없는 값은 추측하지 않고 `{원본코드}(참조표 미등재)`로 남긴다.""",
    "us": """\
`mst/api/openapi_field_foren-us.mst`(UTF-8, 헤더 행 포함)는 위 ✅ 사용 필드를
아래 컬럼으로 변환해 담는다:

| 컬럼 | 원본 순번 | 변환 |
|------|-----------|------|
| 거래소코드 | 1 | **원본 그대로(NAS/NYS/AMX)** — buy/sell/srch가 KB API `krx_cd`로 그대로 전달하므로 번역 금지 |
| 거래소명 | 1 | 표시용 한글명 (NAS→나스닥, NYS→뉴욕, AMX→아멕스) |
| 종목코드 | 2 | 원본 그대로 (티커) |
| 종목명_한글 | 3 | 원본 그대로 |
| 종목명_영문 | 4 | 원본 그대로 |
| 통화코드 | 6 | 원본 그대로 (USD) |
| 종목타입 | 10 | 1→주식, 2→DR, 3→미국워런트, 4→미국우선주, 7→ETF/ETN, 22→워런트, 34→지수 |
| 매매구분 | 12 | 0→제한없음, 1→매수불가(SELL ONLY), 2→매매불가 |
| 매수거래단위 | 14 | 수치→`n주` |
| 매도거래단위 | 15 | 수치→`n주` |
| 소수점매매가능여부 | 33 | 1→소수점매매가능, 2→소수점매매불가 |

코드표에 없는 값은 추측하지 않고 `{원본코드}(참조표 미등재)`로 남긴다.
매매구분 `0`은 공식 코드표(1/2만 정의)에 없지만 "SELL ONLY 구분"이라는 필드
성격상 제한 없음으로 확정했다.""",
}


def write_openapi_md(key, official):
    market = MARKETS[key]
    used = [i for i in CURATION[key] if i["cat"] == "사용"]
    refs = [i for i in CURATION[key] if i["cat"] != "사용"]

    lines = [
        f"# {market['title']}",
        "",
        "KB증권 OpenAPI에서 실제 사용하는 종목마스터 필드 선별표입니다.",
        "",
        "> 자동 생성 문서 — `uv run python -m manage.generate.generate_mst` 재실행으로만 갱신하세요.",
        f"> 필드 순번/코드표의 근거는 공식 명세 `docs/mst/xlsx/{market['source_xlsx']}`입니다.",
        "",
        "## 파일 정보",
        "",
        "| 항목 | 값 |",
        "|------|-----|",
        f"| 원본 파일 | `mst/origin/{market['origin_mst']}` |",
        "| 원본 인코딩 | UTF-8 (구분자 `\\|`) |",
        f"| 공식 명세 필드 수 | {len(official)}개 |",
        f"| 선별 필드 수 | {len(used) + len(refs)}개 (✅ 사용 {len(used)} / 🔶 참조용 {len(refs)}) |",
        "",
        "## 필드 목록",
        "",
        "| 순번 | 구분 | 필드명 | 설명(공식 코드표) | 용도 |",
        "|------|------|--------|-------------------|------|",
    ]
    for item in sorted(CURATION[key], key=lambda x: (x["cat"] != "사용", x["seq"])):
        f = official[item["seq"]]
        cat = "✅ 사용" if item["cat"] == "사용" else "🔶 참조용"
        note = (f["note"] or "-").replace("|", "\\|")
        lines.append(f"| {item['seq']} | {cat} | {f['name']} | {note} | {item['usage']} |")

    lines += [
        "",
        "## openapi_field 변환 규칙",
        "",
        FIELD_MST_NOTES[key],
        "",
        "## 범례",
        "",
        "| 구분 | 설명 |",
        "|------|------|",
        "| ✅ 사용 | 런타임 파일(openapi_field_*.mst)에 실리는 필드 — 주문/시세/검색에 직접 사용 |",
        "| 🔶 참조용 | 필터링 등에 활용 가능하나 런타임 파일에는 싣지 않는 필드 |",
        "",
        "## 관련 문서",
        "",
        f"- 공식 필드 명세(전체): `docs/mst/xlsx/{market['source_xlsx']}`",
        f"- 선별표(엑셀): `docs/mst/xlsx/{market['out_xlsx']}`",
        "- API 목록: `docs/api/api-list.md`",
        "- 생성 스크립트: `manage/generate/generate_mst.py`",
        "",
    ]
    path = MD_DIR / market["out_md"]
    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"생성됨: {path}")


# ── 3단계: mst/api/openapi_field_*.mst (런타임 데이터) ───────────────────────

# 아래 인덱스(0-based)는 공식 명세 순번-1. validate_curation이 순번 실존을 보장한다.


def build_domestic_rows():
    kospi = read_origin_mst("mtsjname.mst", min_cols=27)  # 최대 사용 인덱스 26(순번26)+1
    kosdaq = read_origin_mst("mtsoutjname.mst", min_cols=21)
    rows = []
    for f in kospi:
        rows.append(
            [
                "KOSPI",
                f[0],
                f[1],
                _label(f[3], SEC_GROUP),  # 순번4 증권그룹ID
                _yn(f[20], "관리종목", "정상"),  # 순번21 관리종목여부
                _yn(f[14], "거래정지", "정상"),  # 순번15 거래정지여부
                _unit(f[10]),  # 순번11 매수주문단위
                _yn(f[24], "소수점매매가능", "소수점매매불가"),  # 순번25
                _label(f[25], DECIMAL_STATE),  # 순번26 소수점매매상태
            ]
        )
    for f in kosdaq:
        rows.append(
            [
                "KOSDAQ",
                f[0],
                f[1],
                "",  # 코스닥 원본에 증권그룹ID 없음
                _yn(f[13], "관리종목", "정상"),  # 순번14 관리종목여부
                _yn(f[18], "거래정지", "정상"),  # 순번19 거래정지여부
                "",  # 코스닥 원본에 주문단위 없음
                _yn(f[19], "소수점매매가능", "소수점매매불가"),  # 순번20
                _label(f[20], DECIMAL_STATE),  # 순번21 소수점매매상태
            ]
        )
    return rows, len(kospi), len(kosdaq)


def build_us_rows():
    us = read_origin_mst("FORENMST_US.MST", min_cols=33)  # 최대 사용 인덱스 32(순번33)+1
    rows = []
    for f in us:
        rows.append(
            [
                f[0],
                _label(f[0], EXCHANGE_NAME),  # 순번1 거래소코드(원본 유지) + 표시용
                f[1],
                f[2],
                f[3],
                f[5],  # 순번2 티커 / 3 한글명 / 4 영문명 / 6 통화
                _label(f[9], US_STOCK_TYPE),  # 순번10 종목타입
                _label(f[11], US_TRADE_RESTRICTION),  # 순번12 매매구분
                _unit(f[13]),
                _unit(f[14]),  # 순번14/15 매수/매도거래단위
                _label(f[32], US_DECIMAL_YN),  # 순번33 소수점매매대상종목여부
            ]
        )
    return rows


DOMESTIC_HEADER = [
    "시장구분",
    "종목코드",
    "종목명",
    "종목구분",
    "관리종목여부",
    "거래정지여부",
    "매수주문단위",
    "소수점매매가능여부",
    "소수점매매상태",
]
US_HEADER = [
    "거래소코드",
    "거래소명",
    "종목코드",
    "종목명_한글",
    "종목명_영문",
    "통화코드",
    "종목타입",
    "매매구분",
    "매수거래단위",
    "매도거래단위",
    "소수점매매가능여부",
]


def write_field_mst(filename, header, rows):
    path = OUT_DIR / filename
    with open(path, "w", encoding="utf-8") as f:
        f.write("|".join(header) + "\n")
        f.write("\n".join("|".join(r) for r in rows) + "\n")
    print(f"생성됨: {path}  ({len(rows)}건)")


def sanity_check(domestic_rows, us_rows):
    """생성 결과가 상식적인지 최소 확인 — 실패 시 파일을 이미 썼더라도 명확히 알린다."""
    dom_by_code = {r[1]: r for r in domestic_rows}
    if "005930" not in dom_by_code:
        raise RuntimeError("검증 실패: 삼성전자(005930)가 국내 결과에 없습니다")
    if dom_by_code["005930"][3] != "주식":
        raise RuntimeError(f"검증 실패: 삼성전자 종목구분이 '주식'이 아님: {dom_by_code['005930'][3]}")
    us_by_ticker = {r[2]: r for r in us_rows}
    if "AAPL" not in us_by_ticker:
        raise RuntimeError("검증 실패: AAPL이 해외 결과에 없습니다")
    if us_by_ticker["AAPL"][6] != "주식":
        raise RuntimeError(f"검증 실패: AAPL 종목타입이 '주식'이 아님: {us_by_ticker['AAPL'][6]}")
    unknown = sum(1 for r in domestic_rows + us_rows for c in r if c.endswith("(참조표 미등재)"))
    print(f"검증 통과 (참조표 미등재 값 {unknown}건 — 0이 아니면 코드표 확장 검토)")


def main():
    api_codes = read_api_codes()
    officials = validate_curation(api_codes)

    for key in MARKETS:
        write_openapi_xlsx(key, officials[key])
        write_openapi_md(key, officials[key])

    domestic_rows, n_kospi, n_kosdaq = build_domestic_rows()
    us_rows = build_us_rows()
    write_field_mst("openapi_field_kospi-kosdaq.mst", DOMESTIC_HEADER, domestic_rows)
    write_field_mst("openapi_field_foren-us.mst", US_HEADER, us_rows)
    sanity_check(domestic_rows, us_rows)
    print(f"\n완료: KOSPI {n_kospi} + KOSDAQ {n_kosdaq} + 미국 {len(us_rows)}종목")


if __name__ == "__main__":
    main()
